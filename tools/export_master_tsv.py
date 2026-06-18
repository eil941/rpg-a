# export_master_tsv.py
# master_data.xlsx の各シートを data/master/*.tsv に一括出力するスクリプトです。
#
# 使い方:
#   1. このファイルを rpg-a/tools/export_master_tsv.py に置く
#   2. master_data.xlsx を rpg-a/master_data.xlsx に置く
#   3. コマンドプロンプト/PowerShellで rpg-a フォルダに移動
#   4. python tools/export_master_tsv.py
#
# 必要:
#   pip install openpyxl

from __future__ import annotations

import csv
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


# 「実際に .tres を置き換えたもの」だけを出力対象にする
SHEET_TO_TSV = {
    "item_categories": "item_categories.tsv",
    "items": "items.tsv",
    "equipment": "equipment.tsv",
    "item_effects": "item_effects.tsv",
    "item_effect_links": "item_effect_links.tsv",
    "unit_races": "unit_races.tsv",
    "unit_factions": "unit_factions.tsv",
    "faction_relations": "faction_relations.tsv",
    "element_types": "element_types.tsv",
    "damage_types": "damage_types.tsv",
    "status_effect_types": "status_effect_types.tsv",
    "quests": "quests.tsv",
    "spawn_rules": "spawn_rules.tsv",
    "enemies": "enemies.tsv",
    "npcs": "npcs.tsv",
    "enchantments": "enchantments.tsv",
    "dungeon_spawn_rules": "dungeon_spawn_rules.tsv",
    "unit_spawn_rules": "unit_spawn_rules.tsv",
}


def value_to_text(value: Any) -> str:
    """Excelのセル値をTSV用の文字列に変換する。"""
    if value is None:
        return ""

    if isinstance(value, bool):
        return "true" if value else "false"

    if isinstance(value, int):
        return str(value)

    if isinstance(value, float):
        # 1.0 を 1 として出力する
        if value.is_integer():
            return str(int(value))
        return str(value)

    text = str(value)

    # TSVなのでセル内改行とタブは避ける
    text = text.replace("\r\n", "\\n")
    text = text.replace("\n", "\\n")
    text = text.replace("\r", "\\n")
    text = text.replace("\t", " ")

    return text


def trim_empty_rows(rows: list[list[str]]) -> list[list[str]]:
    """末尾の完全な空行だけ削除する。途中の空欄は保持する。"""
    while rows and all(cell == "" for cell in rows[-1]):
        rows.pop()
    return rows


def trim_empty_columns(rows: list[list[str]]) -> list[list[str]]:
    """
    末尾の完全な空列だけ削除する。
    途中の空列は列ずれ防止のため保持する。
    """
    if not rows:
        return rows

    max_cols = max(len(row) for row in rows)
    keep_cols = max_cols

    for col in range(max_cols - 1, -1, -1):
        has_value = False
        for row in rows:
            if col < len(row) and row[col] != "":
                has_value = True
                break
        if has_value:
            keep_cols = col + 1
            break
    else:
        keep_cols = 0

    return [row[:keep_cols] for row in rows]


def worksheet_to_rows(ws) -> list[list[str]]:
    rows: list[list[str]] = []

    for row in ws.iter_rows():
        rows.append([value_to_text(cell.value) for cell in row])

    rows = trim_empty_rows(rows)
    rows = trim_empty_columns(rows)

    return rows


def export_tsv(rows: list[list[str]], out_path: Path) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)

    with out_path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.writer(
            f,
            delimiter="\t",
            lineterminator="\n",
            quoting=csv.QUOTE_MINIMAL,
        )
        for row in rows:
            writer.writerow(row)


def main() -> int:
    script_path = Path(__file__).resolve()

    # 通常配置:
    # rpg-a/tools/export_master_tsv.py
    # rpg-a/master_data.xlsx
    # rpg-a/data/master/*.tsv
    if script_path.parent.name == "tools":
        project_root = script_path.parent.parent
    else:
        project_root = script_path.parent

    xlsx_path = project_root / "master_data.xlsx"
    out_dir = project_root / "data" / "master"

    if not xlsx_path.exists():
        print(f"[ERROR] master_data.xlsx が見つかりません: {xlsx_path}")
        return 1

    wb = load_workbook(xlsx_path, data_only=True)

    missing_sheets: list[str] = []
    exported_count = 0

    for sheet_name, tsv_name in SHEET_TO_TSV.items():
        if sheet_name not in wb.sheetnames:
            missing_sheets.append(sheet_name)
            continue

        ws = wb[sheet_name]
        rows = worksheet_to_rows(ws)

        out_path = out_dir / tsv_name
        export_tsv(rows, out_path)
        exported_count += 1

        data_rows = max(len(rows) - 1, 0) if rows else 0
        print(f"[OK] {sheet_name} -> {out_path}  rows={data_rows}")

    if missing_sheets:
        print()
        print("[WARN] 見つからなかったシートがあります。出力しませんでした。")
        for name in missing_sheets:
            print(f"  - {name}")

    print()
    print(f"完了: {exported_count} 個のTSVを出力しました。")
    return 0


if __name__ == "__main__":
    sys.exit(main())
